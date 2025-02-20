from cell import *
import openpyxl

def analyzeGroup_RawCentered(cellListRaw, cellList4T1):
    cellList = [[], [], [], [], []]
    group = [[0, 50], [50, 100], [100, 150], [150, 200], [200, 10000]]
    for i in range(len(cellListRaw)):
        cellRaw = cellListRaw[i]
        row = [cellRaw, []]
        for j in range(len(cellList4T1)):
            cell4T1 = cellList4T1[j]
            dist = cellDist(cellRaw, cell4T1)
            for k in range(len(group)):
                if (group[k][0] < dist <= group[k][1]):
                    row[1].append(cell4T1)
    


wb = openpyxl.Workbook()
wb.active.title = "graph"
w1 = wb['graph']

runCount = 0
for i in range(len(headerList)):
    for j in range(len(hourList)):
        for k in range(folderCount[i][j]):
            offset = 7 * (k + 3 * (j + 3 * i)) + 1
            runCount += 1
            print("Trial #{} || {}_{}_{}".format(runCount, headerList[i], hourList[j], str(k+1)), end=" || ")

            extractedCells = extractCells(headerList[i], hourList[j], str(k+1))

            cellListRaw = extractedCells[0]
            cellList4T1 = extractedCells[1]

            stdevAnalysis = analyzeSTDev_RawCentered(cellListRaw, cellList4T1)
            print(stdevAnalysis)
            print()
            

            dataCount = 0
            for m in range(len(stdevAnalysis)):
                data = stdevAnalysis[m][0]
                w1.cell(dataCount + 5, offset) .value= data[0][1].label
                w1.cell(dataCount + 5, offset+1) .value= float(data[0][1].intensity)
                w1.cell(dataCount + 5, offset+2) .value= float(data[0][1].intensityPerArea())
                for n in range(len(data)):
                    w1.cell(dataCount + 5, offset+3) .value= data[n][2].label
                    w1.cell(dataCount + 5, offset+4).value = float(data[n][2].intensity)
                    w1.cell(dataCount + 5, offset+5).value = float(data[n][2].intensityPerArea())
                    w1.cell(dataCount + 5, offset+6).value = float(data[n][0])
                    dataCount += 1

            

filename = "results/group_raw.xlsx"
wb.save(filename)