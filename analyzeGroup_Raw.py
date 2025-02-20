from cell import *
import openpyxl

groupList = [[0, 50], [50, 100], [100, 150], [150, 200], [200, 250], [250, 1000]]

def sortGroupAnalysis_Intensity(data):
    sortedArray = []
    for i in range(len(data)):
        if len(data) != 0 and len(data[i][0]) != 0:
            # print(data[i])
            intensity = data[i][0][0][1].intensity # Raw cell intensity
            sortedArrayLength = len(sortedArray)
            for j in range(len(sortedArray)):
                if (float(sortedArray[j][0][0][1].intensity) > float(intensity)):
                    continue
                else:
                    sortedArray.insert(j, data[i])
                    break
            if len(sortedArray) == sortedArrayLength:
                sortedArray.append(data[i])

    return sortedArray


def analyzeGroup_Raw(cellListRaw, cellList4T1, group):
    thresholdDistanceStart = group[0]
    thresholdDistanceEnd = group[1]
    data = []
    for i in range(len(cellListRaw)):
        distList = []
        effectCellList = []
        for j in range(len(cellList4T1)):
            dist = cellDist(cellListRaw[i], cellList4T1[j])
            distList.append([dist, cellListRaw[i], cellList4T1[j]])

        distSum, distCount = 0, 0
        for k in range(len(distList)):
            if (thresholdDistanceStart <= distList[k][0] <= thresholdDistanceEnd):
                effectCellList.append(distList[k])
                distSum += distList[k][0]
                distCount += 1
        if distCount != 0:
            avgDist = distSum / distCount
        else:
            avgDist = 0

        stdevSum = 0
        if distCount == 1:
            stdev = 0
        else:
            for m in range(len(effectCellList)):
                stdevSum += (avgDist - effectCellList[m][0])**2
            stdev = (stdevSum / (distCount - 1)) ** 0.5
        
        data.append([effectCellList, avgDist, stdev])
    
    return data


for p in range(len(groupList)):
    wb = openpyxl.Workbook()
    wb.active.title = "graph"
    w1 = wb['graph']
    runCount = 0
    for i in range(len(headerList)):
        for j in range(len(hourList)):
            for k in range(folderCount[i][j]):
                offset = 9 * (k + 3 * (j + 3 * i)) + 1
                runCount += 1
                print("Trial #{} || {}_{}_{}".format(runCount, headerList[i], hourList[j], str(k+1)))

                extractedCells = extractCells(headerList[i], hourList[j], str(k+1))

                cellListRaw = extractedCells[0]
                cellList4T1 = extractedCells[1]

                groupAnalysis = analyzeGroup_Raw(cellListRaw, cellList4T1, groupList[p])
                groupAnalysis = sortGroupAnalysis_Intensity(groupAnalysis)

                # print(groupAnalysis)
                # print()
                dataCount = 0
                for m in range(len(groupAnalysis)):
                    data = groupAnalysis[m][0]
                    if len(data) != 0:
                        w1.cell(dataCount + 5, offset+7).value = float(groupAnalysis[m][1])
                        w1.cell(dataCount + 5, offset+8).value = float(groupAnalysis[m][2])
                        w1.cell(dataCount + 5, offset) .value= data[0][1].label
                        w1.cell(dataCount + 5, offset+1) .value= float(data[0][1].intensity)
                        w1.cell(dataCount + 5, offset+2) .value= float(data[0][1].area)
                        for n in range(len(data)):
                            w1.cell(dataCount + 5, offset+3) .value= data[n][2].label
                            w1.cell(dataCount + 5, offset+4).value = float(data[n][2].intensity)
                            w1.cell(dataCount + 5, offset+5).value = float(data[n][2].area)
                            w1.cell(dataCount + 5, offset+6).value = float(data[n][0])
                            dataCount += 1

    filename = "results/group/raw/group_raw_result_" + str(groupList[p][0]) + "_" + str(groupList[p][1]) + ".xlsx"
    wb.save(filename)