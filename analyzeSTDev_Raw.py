from cell import *
import openpyxl


def analyzeSTDev_RawCentered(cellListRaw, cellList4T1):
    threshold = 20 # micrometers
    data = []
    for i in range(len(cellListRaw)):
        distList = []
        effectCellList = []
        minDist = 100000
        for j in range(len(cellList4T1)):
            dist = cellDist(cellListRaw[i], cellList4T1[j])
            if (dist < minDist):
                minDist = dist
            distList.append([dist, cellListRaw[i], cellList4T1[j]])

        distSum, distCount = 0, 0
        for k in range(len(distList)):
            if (distList[k][0] <= minDist + threshold):
                effectCellList.append(distList[k])
                distSum += distList[k][0]
                distCount += 1
        avgDist = distSum / distCount

        stdevSum = 0
        if distCount == 1:
            stdev = 0
        else:
            for m in range(len(effectCellList)):
                stdevSum += (avgDist - effectCellList[m][0])**2
            stdev = (stdevSum / (distCount - 1)) ** 0.5
        
        data.append([effectCellList, avgDist, stdev])
    
    return data



headerList = ["4T1_Raw_1_1", "4T1_Raw_1_2", "4T1_Raw_1_4", "Raw_4T1_1_2", "Raw_4T1_1_4"]
hourList = ["0h", "24h", "72h"]
folderCount = [[3, 3, 2], [3, 3, 3], [1, 3, 3], [3, 2, 2], [3, 3, 2]]

wb = openpyxl.Workbook()
wb.active.title = "graph"
w1 = wb['graph']

runCount = 0
for i in range(len(headerList)):
    for j in range(len(hourList)):
        for k in range(folderCount[i][j]):
            offset = 9 * (k + 3 * (j + 3 * i)) + 1
            runCount += 1
            print("Trial #{} || {}_{}_{}".format(runCount, headerList[i], hourList[j], str(k+1)), end=" || ")

            extractedCells = extractCells(headerList[i], hourList[j], str(k+1))

            cellListRaw = extractedCells[0]
            cellList4T1 = extractedCells[1]

            stdevAnalysis = analyzeSTDev_RawCentered(cellListRaw, cellList4T1)
            print(stdevAnalysis)
            print()
            

            dataCount = 0
            for m in range(len(stdevAnalysis)):
                data = stdevAnalysis[m][0]
                w1.cell(dataCount + 5, offset+7).value = float(stdevAnalysis[m][1])
                w1.cell(dataCount + 5, offset+8).value = float(stdevAnalysis[m][2])
                w1.cell(dataCount + 5, offset) .value= data[0][1].label
                w1.cell(dataCount + 5, offset+1) .value= float(data[0][1].intensity)
                w1.cell(dataCount + 5, offset+2) .value= float(data[0][1].intensityPerArea())
                for n in range(len(data)):
                    w1.cell(dataCount + 5, offset+3) .value= data[n][2].label
                    w1.cell(dataCount + 5, offset+4).value = float(data[n][2].intensity)
                    w1.cell(dataCount + 5, offset+5).value = float(data[n][2].intensityPerArea())
                    w1.cell(dataCount + 5, offset+6).value = float(data[n][0])
                    dataCount += 1

            

filename = "results/stdev_result.xlsx"
wb.save(filename)
